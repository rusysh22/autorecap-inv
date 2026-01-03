import os
import io
import pandas as pd
from flask import Blueprint, render_template, request, jsonify, send_file, current_app
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, date
import re
from . import create_invoice_bp

@create_invoice_bp.route('/')
def index():
    return render_template('create_invoice_index.html')

def safe_float(val):
    """Safely convert value to float, handling Indonesian formats."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    
    s_val = str(val).strip().replace("Rp", "").replace(" ", "")
    if not s_val:
        return 0.0
        
    # Handle Indonesia thousands separator (.) and decimal (,)
    if "," in s_val and "." in s_val:
        if s_val.rfind(',') > s_val.rfind('.'): # 1.234,56
             s_val = s_val.replace('.', '').replace(',', '.')
        else:
             s_val = s_val.replace(',', '')
    elif "." in s_val: # Check if multiple dots (thousands)
         if s_val.count(".") > 1:
             s_val = s_val.replace(".", "")
         else:
             # Ambiguous: 1.000 vs 1.5
             # Assumption: If 3 decimals, likely thousands. If <3, likely decimal.
             parts = s_val.split(".")
             if len(parts[1]) == 3:
                 s_val = s_val.replace(".", "")
             # Else assume standard float
    elif "," in s_val:
        s_val = s_val.replace(",", ".")
        
    try:
        return float(s_val)
    except:
        return 0.0

def terbilang(n):
    """Convert number to Indonesian text."""
    satuan = ["", "SATU", "DUA", "TIGA", "EMPAT", "LIMA", "ENAM", "TUJUH", "DELAPAN", "SEMBILAN", "SEPULUH", "SEBELAS"]
    n = int(n)
    if n >= 0 and n <= 11:
        return " " + satuan[n]
    elif n < 20:
        return terbilang(n % 10) + " BELAS"
    elif n < 100:
        return terbilang(n / 10) + " PULUH" + terbilang(n % 10)
    elif n < 200:
        return " SERATUS" + terbilang(n - 100)
    elif n < 1000:
        return terbilang(n / 100) + " RATUS" + terbilang(n % 100)
    elif n < 2000:
        return " SERIBU" + terbilang(n - 1000)
    elif n < 1000000:
        return terbilang(n / 1000) + " RIBU" + terbilang(n % 1000)
    elif n < 1000000000:
        return terbilang(n / 1000000) + " JUTA" + terbilang(n % 1000000)
    elif n < 1000000000000:
        return terbilang(n / 1000000000) + " MILYAR" + terbilang(n % 1000000000)
    else:
        return ""

@create_invoice_bp.route('/process', methods=['POST'])
def process_files():
    if 'files' not in request.files:
        return jsonify({"error": "No files uploaded"}), 400
    
    files = request.files.getlist('files')
    tax_mode = request.form.get('tax_mode', 'with_tax') # 'with_tax' or 'no_tax'
    
    all_data = []
    anomalies = []
    
    # Target Columns (Based on Image)
    # Col indices (0-based):
    # D = 3 (No. Surat Jalan)
    # H = 7 (PlatNo)
    # I = 8 (Jenis Mobil)
    # U = 20 (PPN)
    # V = 21 (PPH)
    # W = 22 (Total)
    
    for file in files:
        filename = file.filename
        try:
            # Load with header at row 4 (index 3)
            df = pd.read_excel(file, header=3, engine='openpyxl')
            
            # Basic validation
            if df.shape[1] < 23:
                anomalies.append(f"File {filename}: Invalid format (columns missing).")
                continue
                
            # Iterate and clean
            for idx, row in df.iterrows():
                row_num = idx + 5 # 1-based, adjusted for header
                
                # Check for Footer/Signature/Subtotal
                col_a = str(row.iloc[0]).lower() if pd.notna(row.iloc[0]) else ""
                col_d = str(row.iloc[3]).lower() if pd.notna(row.iloc[3]) else ""
                
                # Exclusion keywords
                exclude_keywords = ['total', 'dibuatkan', 'disetujui', 'sub total', 'grand total', 'manager', 'finance']
                if any(k in col_a for k in exclude_keywords) or any(k in col_d for k in exclude_keywords):
                    continue
                
                # Valid Row Logic: Must have No. Surat Jalan (Col D)
                surat_jalan = row.iloc[3]
                if pd.isna(surat_jalan) or str(surat_jalan).strip() == "":
                    continue
                    
                # Extract Data
                # Date Processing: Col K (Index 10) "Waktu Berangkat"
                raw_date = row.iloc[10]
                fmt_date = ""
                try:
                    if pd.notna(raw_date):
                        # If datetime object
                        if isinstance(raw_date, (pd.Timestamp, datetime, date)):
                             fmt_date = raw_date.strftime("%d/%m/%Y")
                        else:
                             # String parsing if needed, e.g. "23/12/2025 0:58"
                             # For now, keep as string or try basic parse
                             s_date = str(raw_date).split()[0] # Take date part
                             fmt_date = s_date
                except:
                    fmt_date = str(raw_date)

                item = {
                    "source_file": filename,
                    "surat_jalan": str(surat_jalan).strip(),
                    "plat_nomor": str(row.iloc[7]).strip() if pd.notna(row.iloc[7]) else "",
                    "jenis_mobil": str(row.iloc[8]).strip().upper() if pd.notna(row.iloc[8]) else "",
                    "rute": str(row.iloc[6]).strip() if pd.notna(row.iloc[6]) else "", # Col G
                    "trip_type": str(row.iloc[9]).strip().upper() if pd.notna(row.iloc[9]) else "", # Col J
                    "date": fmt_date,
                    "dpp": safe_float(row.iloc[15]) if df.shape[1] > 15 else 0,
                    "base_amount_raw": safe_float(row.iloc[15]), # Col P (Changed from O based on user feedback)
                }
                
                raw_ppn = safe_float(row.iloc[20])
                raw_pph = safe_float(row.iloc[21])
                raw_total = safe_float(row.iloc[22])
                
                # Apply Tax Mode Logic
                if tax_mode == 'no_tax':
                    item['ppn'] = 0
                    item['pph'] = 0
                    # Recalculate Total: Start with Total, subtract tax? 
                    # Or is Total Column W inclusive? 
                    # Formula in header: "Total setelah tax".
                    # If No Tax, we should probably take the Pre-Tax value?
                    # Col P (Tarif Sistem) seems to be the base. Or Col O (Harga sebulum tax)?
                    # User said Col P is correct base.
                    base_amount = safe_float(row.iloc[15])
                    item['base_amount'] = base_amount
                    item['final_total'] = base_amount # No Tax
                else:
                    item['ppn'] = raw_ppn
                    item['pph'] = raw_pph
                    item['base_amount'] = safe_float(row.iloc[15])
                    item['final_total'] = raw_total
                
                # Anomaly Checks
                # 1. Jenis Mobil
                if item['jenis_mobil'] not in ['CDDL', 'TWB']:
                     anomalies.append(f"File {filename} Row {row_num}: Jenis Mobil '{item['jenis_mobil']}' invalid (Expected CDDL/TWB).")
                
                all_data.append(item)
                
        except Exception as e:
            anomalies.append(f"File {filename}: Error processing ({str(e)})")

    return jsonify({
        "data": all_data,
        "anomalies": anomalies,
        "count": len(all_data)
    })

@create_invoice_bp.route('/export', methods=['POST'])
def export_excel():
    req_data = request.json
    data = req_data.get('data', [])
    config = req_data.get('config', {}) # {bill_to, ship_to, inv_no, inv_date, due_date, bank_info, currency, tax_rate}
    tax_mode = config.get('tax_mode', 'with_tax') # Default to with_tax or no_tax based on pref, defaulting safely
    
    if not data:
        return jsonify({"error": "No data"}), 400
        
    def format_date_indo(date_str):
        if not date_str: return ""
        try:
            # HTML input date is YYYY-MM-DD
            d = datetime.strptime(date_str, "%Y-%m-%d")
            return d.strftime("%d/%m/%Y")
        except:
            return date_str
            
    inv_date_fmt = format_date_indo(config.get('invoice_date', ''))
    due_date_fmt = format_date_indo(config.get('due_date', ''))
    
    output = io.BytesIO()
    
    # Create Workbook
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Styles
        orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        white_font_bold = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        black_font_bold = Font(color="000000", bold=True, name="Calibri", size=11)
        title_font = Font(color="000000", bold=True, name="Calibri", size=11)
        subtitle_font = Font(color="000000", bold=True, name="Calibri", size=11)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # --- PREPARE DATA ---
        rincian_data = []
        all_dates = []
        
        for idx, item in enumerate(data):
            trip_map = item.get('trip_type', '').upper()
            if 'SEPIHAK' in trip_map: trip_map = 'SINGLE TRIP'
            
            price = item.get('base_amount_raw', 0)
            
            # Collect dates for range
            d_str = item.get('date', '')
            if d_str:
                try:
                    # Try parse DD/MM/YYYY
                    dt = datetime.strptime(d_str, "%d/%m/%Y")
                    all_dates.append(dt)
                except:
                    pass

            rincian_data.append({
                "NO": idx + 1,
                "TANGGAL": d_str,
                "RUTE": item.get('rute', ''),
                "KODE TUGAS": item.get('surat_jalan', ''),
                "TIPE UNIT": item.get('jenis_mobil', ''),
                "TRIP": trip_map,
                "NO POLISI": item.get('plat_nomor', ''),
                "HARGA": price
            })
            
        df_rk = pd.DataFrame(rincian_data)
        
        # Determine Period String
        period_str = "Periode -"
        if all_dates:
            min_date = min(all_dates)
            max_date = max(all_dates)
            # Format: 15-21 Desember 2025
            # Basic mapping for Month if needed, or just use English for now
            # To match screenshot "Desember", we might need a locale map, but let's stick to standard format for now
            period_str = f"Periode {min_date.strftime('%d %B %Y')} - {max_date.strftime('%d %B %Y')}"
            
            # Simplified "15-21 Month Year" if same month/year
            if min_date.month == max_date.month and min_date.year == max_date.year:
                 period_str = f"Periode {min_date.day}-{max_date.day} {min_date.strftime('%B %Y')}"

        # --- SHEET 1: INVOICE ---
        # Ensure INVOICE is the first sheet by using the default 'Sheet' if it exists
        if 'Sheet' in writer.book.sheetnames:
            ws_inv = writer.book['Sheet']
            ws_inv.title = 'INVOICE'
        else:
            ws_inv = writer.book.create_sheet('INVOICE', 0)
            
        writer.sheets['INVOICE'] = ws_inv
        
        # Header Info
        # Header Info
        # Row 2-6: Company Info (Centered C to I)
        # Logo in A2 (No Merge)
        # Logo
        try:
            logo_path = os.path.join(current_app.root_path, 'static', 'create_invoice', 'chijun_sm_f.png')
            if os.path.exists(logo_path):
                img = Image(logo_path)
                
                # Resize keeping aspect ratio
                # Target height matches merge area A2:B5 (4 rows approx 60-80px)
                target_height = 85 
                aspect_ratio = img.width / img.height
                new_width = int(target_height * aspect_ratio)
                
                img.height = target_height
                img.width = new_width
                
                ws_inv.add_image(img, 'A2')
        except Exception as e:
            print(f"Error loading logo: {e}")
        
        # ws_inv['A2'] = "LOGO"  # Removed placeholder

        # Company Text
        ws_inv.merge_cells('C2:H2')
        ws_inv['C2'] = "PT CHIJUN SMART FREIGHT"
        ws_inv['C2'].font = Font(bold=True, size=11)
        ws_inv['C2'].alignment = center_align
        
        ws_inv.merge_cells('C3:H4')
        ws_inv['C3'] = "GEDUNG LANDMARK PLUIT TOWER D2 LT 9 RT.0600 RW.000 PLUIT, PENJARINGAN, KOTA ADM. JAKARTA UTARA, DKI JAKARTA\nPhone: (+62) 821-2459-6308"
        ws_inv['C3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Row 5: NPWP (Shifted up)
        ws_inv.merge_cells('C5:H5')
        ws_inv['C5'] = "NPWP : 0215371659047000"
        ws_inv['C5'].alignment = center_align
        
        # Invoice Title (J-K)
        ws_inv.merge_cells('J3:K5') 
        ws_inv['J3'] = "Invoice"
        ws_inv['J3'].font = Font(bold=True, size=11)
        ws_inv['J3'].alignment = center_align
        
        # --- Bill To / Ship To ---
        # Row 7
        ws_inv['A7'] = "Tagihan Kepada :"
        ws_inv.merge_cells('B7:H7')
        ws_inv['B7'] = config.get('bill_to', '')
        ws_inv['B7'].font = Font(bold=True)
        # Full Border for B7 (Merged B7:H7)
        for c in range(2, 9): # B(2) to H(8)
            ws_inv.cell(row=7, column=c).border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin') if c==2 else None, right=Side(style='thin') if c==8 else None)

        default_address = "Gedung Landmark Pluit Tower, Blok B1 Lantai 8,9A,10A, Jl. Pluit Selatan Raya RT.000, RW.000, Penjaringan\nKota ADM Jakarta Utara, DKI Jakarta"

        ws_inv.merge_cells('B8:H10')
        ws_inv['B8'] = config.get('bill_address', default_address)
        ws_inv['B8'].alignment = Alignment(wrap_text=True, vertical='top')
        # Box Border for Address B8:H10
        for r_addr in range(8, 11):
            for c in range(2, 9):
                 border_style = Border(
                     left=Side(style='thin') if c==2 else None, 
                     right=Side(style='thin') if c==8 else None,
                     bottom=Side(style='thin') if r_addr==10 else None
                 )
                 # Merge existing? No, overwrite safe.
                 ws_inv.cell(row=r_addr, column=c).border = border_style
        
        ws_inv.merge_cells('B11:H11')
        ws_inv['B11'] = "NPWP : " + config.get('bill_npwp', '0735697740041000')
        for c in range(2, 9):
            ws_inv.cell(row=11, column=c).border = Border(bottom=Side(style='thin'), left=Side(style='thin') if c==2 else None, right=Side(style='thin') if c==8 else None)
        
        ws_inv['A12'] = "Dikirim ke :"
        
        ws_inv.merge_cells('B12:H12')
        ws_inv['B12'] = config.get('ship_to', '')
        ws_inv['B12'].font = Font(bold=True)
        # Full Border for B12
        for c in range(2, 9): 
            ws_inv.cell(row=12, column=c).border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin') if c==2 else None, right=Side(style='thin') if c==8 else None)
        
        ws_inv.merge_cells('B13:H15')
        ws_inv['B13'] = config.get('ship_address', default_address)
        ws_inv['B13'].alignment = Alignment(wrap_text=True, vertical='top')
        # Box Border for Address B13:H15
        for r_addr in range(13, 16):
            for c in range(2, 9):
                 border_style = Border(
                     left=Side(style='thin') if c==2 else None, 
                     right=Side(style='thin') if c==8 else None,
                     bottom=Side(style='thin') if r_addr==15 else None
                 )
                 ws_inv.cell(row=r_addr, column=c).border = border_style

        # Row 16: No Merge (User Request)
        # Spacer or specific content? Image shows just border or empty?
        # Image shows borders on B16? No, "merge cell pada baris 16 ... tidak perlu".
        # So leave A16:I16 empty/unmerged? Ok.


        # --- Right Side Details (J-K) ---
        # J7-K7: No. Invoice Header
        ws_inv.merge_cells('J7:K7')
        ws_inv['J7'] = "No. Invoice"
        ws_inv['J7'].alignment = center_align
        ws_inv['J7'].border = thin_border
        ws_inv['K7'].border = thin_border
        
        # J8-K9: Invoice No Value
        ws_inv.merge_cells('J8:K9')
        ws_inv['J8'] = config.get('invoice_no', '')
        ws_inv['J8'].alignment = center_align
        ws_inv['J8'].border = thin_border
        for r in range(8, 10):
            for c in range(10, 12):
                ws_inv.cell(row=r, column=c).border = thin_border
        
        # J10-K10: No. Faktur Header
        ws_inv.merge_cells('J10:K10')
        ws_inv['J10'] = "No. Faktur"
        ws_inv['J10'].alignment = center_align
        # No bottom border for Row 10 (User request)
        no_bottom = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))
        ws_inv['J10'].border = no_bottom
        ws_inv['K10'].border = no_bottom
        
        # J11-K11: No. Faktur Value
        ws_inv.merge_cells('J11:K11')
        ws_inv['J11'] = "" 
        # To visually look like one box with Row 10, Row 11 needs NO Top Border.
        no_top = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))
        ws_inv['J11'].border = no_top
        ws_inv['K11'].border = no_top
        
        # Row 12
        ws_inv['J12'] = "Invoice Date"
        ws_inv['J12'].alignment = center_align
        ws_inv['J12'].border = no_bottom
        ws_inv['K12'] = "Currency"
        ws_inv['K12'].alignment = center_align
        ws_inv['K12'].border = no_bottom
        
        # Row 13
        ws_inv['J13'] = inv_date_fmt
        ws_inv['J13'].alignment = center_align
        ws_inv['J13'].border = no_top
        ws_inv['K13'] = config.get('currency', 'IDR')
        ws_inv['K13'].alignment = center_align
        ws_inv['K13'].border = no_top
        
        # Row 14
        ws_inv['J14'] = "Tax Rate"
        ws_inv['J14'].alignment = center_align
        ws_inv['J14'].border = no_bottom
        ws_inv['K14'] = "Due Date"
        ws_inv['K14'].alignment = center_align
        ws_inv['K14'].border = no_bottom
        
        # Row 15
        ws_inv['J15'] = "" 
        ws_inv['J15'].border = no_top
        ws_inv['K15'] = due_date_fmt
        ws_inv['K15'].alignment = center_align
        ws_inv['K15'].border = no_top
        
        # Row 16 - Spacer? Image shows table starts Row 17 (Header).

        # Table Header (Row 17)
        ws_inv['A17'] = "No."
        ws_inv.merge_cells('B17:E17')
        ws_inv['B17'] = "Deskripsi"
        ws_inv['F17'] = "Tipe Mobil"
        ws_inv['G17'] = "Total Rit"
        ws_inv['H17'] = "Biaya Per Rit"
        ws_inv['I17'] = "Diskon"
        ws_inv['J17'] = "PPN"
        ws_inv['K17'] = "Total Biaya"
        
        for col in ['A', 'B', 'F', 'G', 'H', 'I', 'J', 'K']:
             c = ws_inv[f'{col}17']
             c.fill = orange_fill
             c.font = black_font_bold
             c.alignment = center_align
             c.border = thin_border
             if col == 'B': 
                 for sub_c in range(2, 6): 
                     ws_inv.cell(row=17, column=sub_c).border = thin_border
                     ws_inv.cell(row=17, column=sub_c).fill = orange_fill

        # Aggregation Logic
        inv_rows = []
        if not df_rk.empty:
            grouped = df_rk.groupby(['RUTE', 'TIPE UNIT', 'TRIP']).agg(
                TOTAL_RITASE=('NO', 'count'),
                HARGA_RIT=('HARGA', 'first'),
                TOTAL_HARGA=('HARGA', 'sum')
            ).reset_index().sort_values(by=['RUTE'])
            
            for i, row in grouped.iterrows():
                desc_1 = f"Biaya Transportasi {period_str}" # Line 1
                desc_2 = f"{row['RUTE']}" # Line 2 (Route)
                # desc_3 = "Jasa Angkutan Umum (Plat Kuning)"
                
                # Image Logic:
                # Desc looks like: "Biaya Transportasi Periode 1-7 Desember 2025" (Newline) "CRN-CKP-SMI"
                
                # PPN Calculation (1.1% based on image)
                # Image: DPP 8.460.000, PPN 93.060 (approx 1.1%)
                # 8460000 * 0.011 = 93060. Exact.
                ppn_val = row['TOTAL_HARGA'] * 0.011 if tax_mode == 'with_tax' else 0
                
                inv_rows.append({
                    "no": i + 1,
                    "desc_lines": [desc_1, desc_2],
                    "type": row['TIPE UNIT'],
                    "rit": row['TOTAL_RITASE'],
                    "price_rit": row['HARGA_RIT'],
                    "disc": "-",
                    "ppn": ppn_val,
                    "total": row['TOTAL_HARGA']
                })
        
        start_row = 18
        if not inv_rows:
             inv_rows.append({"no": 1, "desc_lines": ["No Data"], "type": "", "rit": "", "price_rit": 0, "disc": "-", "ppn": 0, "total": 0})

        for item in inv_rows:
            ws_inv[f'A{start_row}'] = item['no']
            ws_inv[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws_inv[f'A{start_row}'].border = thin_border
            
            ws_inv.merge_cells(f'B{start_row}:E{start_row}')
            ws_inv[f'B{start_row}'] = "\n".join(item['desc_lines'])
            ws_inv[f'B{start_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for c in range(2, 6): ws_inv.cell(row=start_row, column=c).border = thin_border
            
            ws_inv[f'F{start_row}'] = item['type']
            ws_inv[f'F{start_row}'].alignment = center_align
            ws_inv[f'F{start_row}'].border = thin_border
            
            ws_inv[f'G{start_row}'] = item['rit']
            ws_inv[f'G{start_row}'].alignment = center_align
            ws_inv[f'G{start_row}'].border = thin_border
            
            # User request: Col H-K Middle Align (Vertical Center) with Accounting Format
            # Accounting Format with Rp: _("Rp"* #,##0_);_("Rp"* (#,##0);_("Rp"* "-"??_);_(@_)
            res_accounting_fmt = '_("Rp"* #,##0_);_("Rp"* (#,##0);_("Rp"* "-"??_);_(@_)'
            align_acc = Alignment(horizontal='right', vertical='center') 
            
            ws_inv[f'H{start_row}'] = item['price_rit']
            ws_inv[f'H{start_row}'].number_format = res_accounting_fmt
            ws_inv[f'H{start_row}'].alignment = align_acc
            ws_inv[f'H{start_row}'].border = thin_border
            
            # Disc (I)
            ws_inv[f'I{start_row}'] = item['disc']
            # If disc is "-", accounting format handles it if it's text/zero? 
            # If text "-", format might treat as text. If 0, it shows "-".
            # The value is item['disc'] which is "-" string. 
            # To use accounting format effectively for '-', value should ideally be 0.
            # But let's stick to what we have. Align center if text?
            # User asked H-K accounting.
            ws_inv[f'I{start_row}'].number_format = res_accounting_fmt
            ws_inv[f'I{start_row}'].alignment = align_acc
            ws_inv[f'I{start_row}'].border = thin_border
            
            ws_inv[f'J{start_row}'] = item['ppn']
            ws_inv[f'J{start_row}'].number_format = res_accounting_fmt
            ws_inv[f'J{start_row}'].alignment = align_acc
            ws_inv[f'J{start_row}'].border = thin_border
            
            ws_inv[f'K{start_row}'] = item['total']
            ws_inv[f'K{start_row}'].number_format = res_accounting_fmt
            ws_inv[f'K{start_row}'].alignment = align_acc
            ws_inv[f'K{start_row}'].border = thin_border
            
            ws_inv.row_dimensions[start_row].height = 30
            start_row += 1

        # Summary Section
        sum_row = start_row
        
        total_dpp = sum(x['total'] for x in inv_rows)
        total_ppn = sum(x['ppn'] for x in inv_rows)
        total_pph = total_dpp * 0.02
        
        # New Tax Logic: Final = DPP + PPN - PPh
        final_payment = total_dpp + total_ppn - total_pph if tax_mode == 'with_tax' else total_dpp
        
        terbilang_txt = terbilang(final_payment).strip() + " RUPIAH"
        
        # Terbilang Box (Rows sum_row, sum_row+1 | Cols A-G)
        ws_inv.merge_cells(f'A{sum_row}:G{sum_row+1}')
        ws_inv[f'A{sum_row}'] = terbilang_txt.upper()
        ws_inv[f'A{sum_row}'].font = Font(bold=True)
        ws_inv[f'A{sum_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for r_tb in range(sum_row, sum_row+2):
            for c_tb in range(1, 8): ws_inv.cell(row=r_tb, column=c_tb).border = thin_border
        
        # --- Totals Section (Right Side H-K) ---
        
        # 1. Total Diskon (Row sum_row)
        ws_inv.merge_cells(f'H{sum_row}:J{sum_row}')
        ws_inv[f'H{sum_row}'] = "Total Diskon"
        ws_inv[f'K{sum_row}'] = "-" # Or 0 if we want accounting dash
        ws_inv[f'K{sum_row}'].number_format = res_accounting_fmt
        ws_inv[f'K{sum_row}'].alignment = align_acc
        
        # 2. Total DPP (Row sum_row+1)
        ws_inv.merge_cells(f'H{sum_row+1}:J{sum_row+1}')
        ws_inv[f'H{sum_row+1}'] = "Total Dasar Pengenaan Pajak"
        ws_inv[f'K{sum_row+1}'] = total_dpp
        ws_inv[f'K{sum_row+1}'].number_format = res_accounting_fmt
        ws_inv[f'K{sum_row+1}'].alignment = align_acc

        # 3. Total PPN (Row sum_row+2)
        ws_inv.merge_cells(f'H{sum_row+2}:J{sum_row+2}')
        ws_inv[f'H{sum_row+2}'] = "Total PPN (1.1%)"
        ws_inv[f'K{sum_row+2}'] = total_ppn
        ws_inv[f'K{sum_row+2}'].number_format = res_accounting_fmt
        ws_inv[f'K{sum_row+2}'].alignment = align_acc
        
        # 4. Total PPh (Row sum_row+3) 
        ws_inv.merge_cells(f'H{sum_row+3}:J{sum_row+3}')
        ws_inv[f'H{sum_row+3}'] = "Total PPh 23 (2%)"
        ws_inv[f'K{sum_row+3}'] = total_pph
        ws_inv[f'K{sum_row+3}'].number_format = res_accounting_fmt
        ws_inv[f'K{sum_row+3}'].alignment = align_acc
        
        # 5. Total Bayar (Row sum_row+4)
        ws_inv.merge_cells(f'H{sum_row+4}:J{sum_row+4}')
        ws_inv[f'H{sum_row+4}'] = "Total Bayar"
        ws_inv[f'K{sum_row+4}'] = final_payment
        ws_inv[f'K{sum_row+4}'].number_format = res_accounting_fmt
        ws_inv[f'K{sum_row+4}'].alignment = align_acc
        
        # Apply Borders to Right Side (H-K, Rows sum_row to sum_row+4)
        for r_tot in range(sum_row, sum_row+5):
            for c_tot in range(8, 12): # H(8) to K(11)
                bs = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                # Handle merged cells borders
                ws_inv.cell(row=r_tot, column=c_tot).border = thin_border
        
        # --- Additional Left Content (A-G) ---
        
        # Combined Deskripsi & Bank Info (Rows sum_row+2 to sum_row+4) matches PPN, PPh, Total Bayar
        # User request: Merge Baris 21-23 Col A-G with Deskripsi + Bank Info
        ws_inv.merge_cells(f'A{sum_row+2}:G{sum_row+4}')
        
        bank_account = config.get('bank_info', 'BCA 1685681899 (KCU PLUIT)')
        combined_text = f"Deskripsi : Biaya Transportasi {period_str}\n" + \
                        "Silahkan melakukan pembayaran dengan melakukan transfer ke rekening :\n" + \
                        "PT CHIJUN SMART FREIGHT\n" + \
                        f"{bank_account}"
                        
        ws_inv[f'A{sum_row+2}'] = combined_text
        ws_inv[f'A{sum_row+2}'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        # User request: Bottom border for Row 23 (sum_row+4) Cols A-G
        # Also Left border for A(sum_row+2) to A(sum_row+4)
        for r_bank in range(sum_row+2, sum_row+5):
             # Left Border for Col A
             ws_inv.cell(row=r_bank, column=1).border = Border(left=Side(style='thin'), bottom=Side(style='thin') if r_bank==sum_row+4 else None)
             
        for c_bank in range(2, 8): # B-G (Bottom border only for last row)
            ws_inv.cell(row=sum_row+4, column=c_bank).border = Border(bottom=Side(style='thin'))
        
        # Signatures
        # Spacer Row (sum_row+5)
        # Headers Row (sum_row+6)
        sig_starting_row = sum_row + 6
        
        # 1. Prepared By (A-B)
        ws_inv.merge_cells(f'A{sig_starting_row}:B{sig_starting_row}')
        ws_inv[f'A{sig_starting_row}'] = "Prepared By"
        ws_inv[f'A{sig_starting_row}'].alignment = Alignment(horizontal='left')
        
        # 2. Checked By (D-E) - Spacer C
        ws_inv.merge_cells(f'D{sig_starting_row}:E{sig_starting_row}')
        ws_inv[f'D{sig_starting_row}'] = "Checked By"
        ws_inv[f'D{sig_starting_row}'].alignment = Alignment(horizontal='left')
        
        # 3. Approved By (G-H) - Spacer F
        ws_inv.merge_cells(f'G{sig_starting_row}:H{sig_starting_row}')
        ws_inv[f'G{sig_starting_row}'] = "Approved By"
        ws_inv[f'G{sig_starting_row}'].alignment = Alignment(horizontal='left')
        
        # 4. Received By (J-K) - Spacer I
        ws_inv.merge_cells(f'J{sig_starting_row}:K{sig_starting_row}')
        ws_inv[f'J{sig_starting_row}'] = "Received By"
        ws_inv[f'J{sig_starting_row}'].alignment = Alignment(horizontal='left')
        
        # Signature Space
        line_row = sig_starting_row + 7
        
        # Draw Bottom Borders for lines
        # Group 1: A-B
        for c in range(1, 3): ws_inv.cell(row=line_row, column=c).border = Border(bottom=Side(style='thin'))
        # Group 2: D-E (4, 5)
        for c in range(4, 6): ws_inv.cell(row=line_row, column=c).border = Border(bottom=Side(style='thin'))
        # Group 3: G-H (7, 8)
        for c in range(7, 9): ws_inv.cell(row=line_row, column=c).border = Border(bottom=Side(style='thin'))
        # Group 4: J-K (10, 11)
        for c in range(10, 12): ws_inv.cell(row=line_row, column=c).border = Border(bottom=Side(style='thin'))
        
        # Date Row (line_row + 1)
        date_row = line_row + 1
        ws_inv[f'A{date_row}'] = "Tanggal:"
        ws_inv[f'D{date_row}'] = "Tanggal:"
        ws_inv[f'G{date_row}'] = "Tanggal:"
        ws_inv[f'J{date_row}'] = "Tanggal:"

        # Column Widths
        ws_inv.column_dimensions['A'].width = 5
        ws_inv.column_dimensions['B'].width = 25
        ws_inv.column_dimensions['C'].width = 10
        ws_inv.column_dimensions['D'].width = 10
        ws_inv.column_dimensions['E'].width = 10
        ws_inv.column_dimensions['F'].width = 15
        ws_inv.column_dimensions['G'].width = 10
        ws_inv.column_dimensions['H'].width = 15
        ws_inv.column_dimensions['I'].width = 15
        ws_inv.column_dimensions['J'].width = 15
        ws_inv.column_dimensions['K'].width = 20
        
        # --- Page Setup & Margins ---
        # User specified: 
        # Paper: A4, Portrait
        # Scaling: Fit to 1 page wide, unknown height (FitAllColumnsOnOnePage equivalent)
        # Margins (CM): Top 1, Left 1, Right 1, Bottom 2.5, Header 1.3, Footer 1.3
        
        ws_inv.page_setup.orientation = ws_inv.ORIENTATION_PORTRAIT
        ws_inv.page_setup.paperSize = ws_inv.PAPERSIZE_A4
        ws_inv.page_setup.fitToWidth = 1
        ws_inv.page_setup.fitToHeight = 0 # Automatic
        
        # Convert CM to Inches (1 inch = 2.54 cm)
        cm_to_inch = 1 / 2.54
        ws_inv.page_margins.top = 1 * cm_to_inch
        ws_inv.page_margins.left = 1 * cm_to_inch
        ws_inv.page_margins.right = 1 * cm_to_inch
        ws_inv.page_margins.bottom = 2.5 * cm_to_inch
        ws_inv.page_margins.header = 1.3 * cm_to_inch
        ws_inv.page_margins.footer = 1.3 * cm_to_inch


        # --- SHEET 2: KWITANSI ---        
        # --- SHEET 2: KWITANSI ---
        ws_kw = writer.book.create_sheet('KWITANSI')
        writer.sheets['KWITANSI'] = ws_kw
        ws_kw.sheet_view.showGridLines = False
        
        # Styles
        kw_font_bold = Font(name='Calibri', size=11, bold=True)
        kw_font_reg = Font(name='Calibri', size=11)
        
        # --- HEADER (Rows 4-8) ---
        # Logo at A4
        try:
            if os.path.exists(logo_path):
                img_kw = Image(logo_path)
                # Keep same size as Invoice or slightly larger
                target_h_kw = 60
                ar_kw = img_kw.width / img_kw.height
                img_kw.height = target_h_kw
                img_kw.width = int(target_h_kw * ar_kw)
                ws_kw.add_image(img_kw, 'A4')
        except:
            pass

        # C4: Company Name
        ws_kw['C4'] = "PT CHIJUN SMART FREIGHT"
        ws_kw['C4'].font = kw_font_bold
        
        # J4: Invoice No (Merged J4:L4 approx to fit)
        # Assuming layout is wide
        ws_kw['J4'] = "NO: " + config.get('invoice_no', '')
        ws_kw['J4'].font = kw_font_bold
        ws_kw['J4'].alignment = Alignment(horizontal='right')
        
        # Row 8: Title "KWITANSI"
        ws_kw.merge_cells('A8:L8') # Center across sheet
        ws_kw['A8'] = "KWITANSI"
        ws_kw['A8'].font = kw_font_bold
        ws_kw['A8'].alignment = center_align
        # Double bottom border for Row 8
        for c_k in range(1, 13): # A-L
            ws_kw.cell(row=8, column=c_k).border = Border(bottom=Side(style='double'))

        # --- BODY (Rows 10-14) ---
        ws_kw['A10'] = "Untuk transaksi tersebut dibawah ini"
        ws_kw['A10'].font = kw_font_reg
        
        # Row 11: Nama Perusahaan
        ws_kw['A11'] = "Nama Perusahaan"
        ws_kw['B11'] = ":"
        ws_kw['C11'] = config.get('bill_to', '')
        ws_kw['C11'].font = kw_font_reg
        
        # Row 12: Jumlah
        ws_kw['A12'] = "Jumlah"
        ws_kw['B12'] = ":"
        ws_kw['C12'] = f"Rp {final_payment:,.0f}".replace(",", ".")
        ws_kw['C12'].font = kw_font_reg
        
        # Row 13: Transaksi
        ws_kw['A13'] = "Transaksi"
        ws_kw['B13'] = ":"
        ws_kw['C13'] = f"Biaya pengiriman ekspedisi logistik {config.get('bill_to', 'PT Global Jet Express')} (termasuk Ppn-Pph 23)"
        ws_kw['C13'].font = kw_font_reg
        
        # --- AMOUNT & TERBILANG (Rows 18-20) ---
        # B18: Rp, C18: Amount (Bold)
        ws_kw['B18'] = "Rp"
        ws_kw['B18'].font = kw_font_bold
        ws_kw['B18'].alignment = Alignment(horizontal='right')
        
        ws_kw['C18'] = f"{final_payment:,.0f}".replace(",", ".")
        ws_kw['C18'].font = kw_font_bold
        
        # B19: Terbilang
        ws_kw['B19'] = "Terbilang"
        ws_kw['B19'].font = kw_font_bold
        ws_kw['B19'].alignment = Alignment(horizontal='right', vertical='top')
        
        # C19: Text (Merged C19:L20)
        ws_kw.merge_cells('C19:L20')
        ws_kw['C19'] = terbilang_txt
        ws_kw['C19'].font = kw_font_reg
        ws_kw['C19'].alignment = Alignment(wrap_text=True, vertical='top')

        # --- DATE & SIGNATURE (Row 21) ---
        # Date at J21 (Right aligned)
        # Format: Jakarta, 10 December 2025
        # Need English month? User image says "December". 
        # For now using simple format or standard English if compatible.
        try:
            d_obj = datetime.strptime(config.get('invoice_date', ''), "%Y-%m-%d")
            # Creating dummy English month list or use %B if locale supports
            eng_months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
            date_str = f"{d_obj.day} {eng_months[d_obj.month-1]} {d_obj.year}"
        except:
             date_str = config.get('invoice_date', '')
             
        ws_kw['I21'] = f"Jakarta, {date_str}"
        ws_kw['I21'].alignment = Alignment(horizontal='center') # Looks somewhat centered in right area
        ws_kw.merge_cells('I21:L21')

        # --- PAYMENT BOXES (Rows 22-24) ---
        ws_kw['A22'] = "Pembayaran dianggap sah"
        ws_kw['A23'] = "setelah cek /giro diuangkan"
        ws_kw['A22'].font = kw_font_reg
        ws_kw['A23'].font = kw_font_reg
        
        # Boxes at C22, C23, C24
        box_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        ws_kw['C22'] = "Cek :"
        ws_kw['C22'].border = box_border
        
        ws_kw['C23'] = "Giro :"
        ws_kw['C23'].border = box_border
        
        ws_kw['C24'] = "Cash :"
        ws_kw['C24'].border = box_border
        
        # --- SIGNATURE (Left) ---
        ws_kw['A29'] = "Mengetahui,"
        ws_kw['A36'] = "PT CHIJUN SMART FREIGHT"
        ws_kw['A36'].font = kw_font_bold
        
        # --- COLUMN WIDTHS ---
        ws_kw.column_dimensions['A'].width = 30
        ws_kw.column_dimensions['B'].width = 12
        ws_kw.column_dimensions['C'].width = 25
        # D-L auto or standard
        for ch in "DEFGHIJKLMNOP":
            ws_kw.column_dimensions[ch].width = 10
        
        # Widths
        ws_kw.column_dimensions['A'].width = 25
        ws_kw.column_dimensions['C'].width = 5
        ws_kw.column_dimensions['D'].width = 30
        
        # --- SHEET 3: RINCIAN RITASE ---
        # Aggregation Logic
        # Group by RUTE, TIPE UNIT, TRIP
        if not df_rk.empty:
            grouped = df_rk.groupby(['RUTE', 'TIPE UNIT', 'TRIP']).agg(
                TOTAL_RITASE=('NO', 'count'),
                HARGA_RIT=('HARGA', 'first'), # Assume same price for same group
                TOTAL_HARGA=('HARGA', 'sum')
            ).reset_index()
            
            # Sort for consistency
            grouped = grouped.sort_values(by=['RUTE'])
            
            # Add NO and TANGGAL columns
            grouped['NO'] = range(1, len(grouped) + 1)
            grouped['TANGGAL'] = period_str.replace("Periode ", "").upper() # Just the date part
            
            # Reorder columns
            # Rename TIPE UNIT -> TYPE, TRIP -> JENIS TRIP to match screenshot headers
            grouped = grouped.rename(columns={
                'TIPE UNIT': 'TYPE', 
                'TRIP': 'JENIS TRIP',
                'TOTAL_RITASE': 'TOTAL RITASE',
                'HARGA_RIT': 'HARGA/RIT',
                'TOTAL_HARGA': 'TOTAL HARGA'
            })
            final_cols = ['NO', 'TANGGAL', 'RUTE', 'TYPE', 'JENIS TRIP', 'TOTAL RITASE', 'HARGA/RIT', 'TOTAL HARGA']
            
            grouped = grouped[final_cols]
        else:
            grouped = pd.DataFrame(columns=['NO', 'TANGGAL', 'RUTE', 'TYPE', 'JENIS TRIP', 'TOTAL RITASE', 'HARGA/RIT', 'TOTAL HARGA'])

        # Write to Excel
        # Start at row 6 (index 5) to follow user screenshot (Row 1-5 Titles, Row 6 Header)
        grouped.to_excel(writer, sheet_name='RINCIAN RITASE', index=False, startrow=5)
        ws_rit = writer.sheets['RINCIAN RITASE']
        
        # Helper for Indonesian Months
        months_id = {
            1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni',
            7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
        }
        
        # Determine Period String for Header
        period_str = "Periode -"
        if all_dates:
            min_date = min(all_dates)
            max_date = max(all_dates)
            
            m_start = months_id[min_date.month]
            m_end = months_id[max_date.month]
            
            if min_date.month == max_date.month and min_date.year == max_date.year:
                 period_str = f"Periode {min_date.day}-{max_date.day} {m_start} {min_date.year}"
            elif min_date.year == max_date.year:
                 period_str = f"Periode {min_date.day} {m_start} - {max_date.day} {m_end} {min_date.year}"
            else:
                 period_str = f"Periode {min_date.day} {m_start} {min_date.strftime('%Y')} - {max_date.day} {m_end} {max_date.strftime('%Y')}"

        # Determine Project Name from Filename Code
        # Default
        project_name = "PROJEK J&T EXPRESS"
        if data:
            first_file = data[0].get('source_file', '')
            match = re.match(r'^.+?_([A-Za-z0-9]+)_', first_file)
            if match:
                city_code = match.group(1).upper()
                project_name = f"PROJEK J&T EXPRESS {city_code}"
            else:
                 # Fallback: try to see if 'J2' logic was meant (e.g. from file content? but file is closed).
                 # Stick to filename or city code from route?
                 # If no code found in filename pattern, maybe check Route?
                 pass

        # Add Titles (Rows 1-5)
        ws_rit.merge_cells('A1:H1')
        ws_rit['A1'] = "RINCIAN RITASE"
        ws_rit['A1'].font = Font(name='Calibri', size=11, bold=True)
        ws_rit['A1'].alignment = center_align

        ws_rit.merge_cells('A2:H2')
        ws_rit['A2'] = project_name
        ws_rit['A2'].font = Font(name='Calibri', size=11, bold=False) # Or bold? Screenshot looks normal/bold? Let's go normal based on prev, or bold if title.
        ws_rit['A2'].alignment = center_align

        ws_rit.merge_cells('A3:H3')
        ws_rit['A3'] = "LAPORAN PEMAKAIAN KENDARAAN"
        ws_rit['A3'].font = Font(name='Calibri', size=11, bold=False)
        ws_rit['A3'].alignment = center_align
        
        ws_rit.merge_cells('A4:H4')
        ws_rit['A4'] = "PT CHIJUN SMART FREIGHT"
        ws_rit['A4'].font = Font(name='Calibri', size=11, bold=False)
        ws_rit['A4'].alignment = center_align

        ws_rit.merge_cells('A5:H5')
        ws_rit['A5'] = period_str
        ws_rit['A5'].font = Font(name='Calibri', size=11, bold=False)
        ws_rit['A5'].alignment = center_align
            
        # Style Table Header (Row 6)
        for cell in ws_rit[6]:
            cell.fill = orange_fill
            cell.font = black_font_bold
            cell.alignment = center_align
            cell.border = thin_border
            
        # Style Data (Starts Row 7)
        rit_data_len = len(grouped)
        rit_total_sum = grouped['TOTAL HARGA'].sum() if not grouped.empty else 0
        
        for row in ws_rit.iter_rows(min_row=7, max_row=7+rit_data_len-1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_align
                if cell.column in [7, 8]: # Harga Rit (G), Total Harga (H)
                    cell.number_format = 'Rp #,##0'

        # Total Row
        tot_row = 7 + rit_data_len
        ws_rit.merge_cells(f'A{tot_row}:G{tot_row}')
        ws_rit[f'A{tot_row}'] = "TOTAL"
        ws_rit[f'A{tot_row}'].fill = orange_fill
        ws_rit[f'A{tot_row}'].font = black_font_bold
        ws_rit[f'A{tot_row}'].alignment = center_align
        
        # Apply border/fill to merged cells
        for c_idx in range(1, 8):
            ws_rit.cell(row=tot_row, column=c_idx).border = thin_border
            ws_rit.cell(row=tot_row, column=c_idx).fill = orange_fill
            
        ws_rit[f'H{tot_row}'] = rit_total_sum
        ws_rit[f'H{tot_row}'].fill = orange_fill
        ws_rit[f'H{tot_row}'].font = black_font_bold
        ws_rit[f'H{tot_row}'].border = thin_border
        ws_rit[f'H{tot_row}'].number_format = 'Rp #,##0'
        
        # Signatures
        sig_rit_start = tot_row + 3
        ws_rit[f'B{sig_rit_start}'] = "Mengetahui"
        ws_rit[f'B{sig_rit_start+6}'] = "PT CHIJUN SMART FREIGHT"
        ws_rit[f'B{sig_rit_start+6}'].font = Font(bold=True)
        
        ws_rit[f'G{sig_rit_start}'] = "Mengetahui"
        ws_rit[f'G{sig_rit_start+6}'] = "PT GLOBAL JET EXPRESS"
        ws_rit[f'G{sig_rit_start+6}'].font = Font(bold=True)

        # Col Widths
        ws_rit.column_dimensions['A'].width = 5
        ws_rit.column_dimensions['B'].width = 25
        ws_rit.column_dimensions['C'].width = 15
        ws_rit.column_dimensions['D'].width = 15
        ws_rit.column_dimensions['E'].width = 15
        ws_rit.column_dimensions['F'].width = 15
        ws_rit.column_dimensions['G'].width = 18
        ws_rit.column_dimensions['H'].width = 18

        # --- SHEET 4: RINCIAN KENDARAAN ---
        # Write Data
        df_rk.to_excel(writer, sheet_name='RINCIAN KENDARAAN', index=False, startrow=0)
        
        ws = writer.sheets['RINCIAN KENDARAAN']
        
        # Apply Header Styles (Row 1)
        for cell in ws[1]:
            cell.fill = orange_fill
            cell.font = black_font_bold
            cell.alignment = center_align
            cell.border = thin_border
            
        # Apply Data Styles & Borders
        for row in ws.iter_rows(min_row=2, max_row=len(rincian_data)+1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_align
                if cell.column == 8: # Col H
                     cell.number_format = '#,##0'
                     
        # Add Total Row
        total_row_idx = len(rincian_data) + 2
        ws.merge_cells(f'A{total_row_idx}:G{total_row_idx}')
        total_label_cell = ws[f'A{total_row_idx}']
        total_label_cell.value = "TOTAL"
        total_label_cell.fill = orange_fill
        total_label_cell.font = black_font_bold
        total_label_cell.alignment = center_align
        total_label_cell.border = thin_border
        
        # Fill merged cells border
        for col in range(1, 8): # A to G
             ws.cell(row=total_row_idx, column=col).border = thin_border
             ws.cell(row=total_row_idx, column=col).fill = orange_fill

        total_val_cell = ws[f'H{total_row_idx}']
        total_val_cell.value = df_rk['HARGA'].sum() if not df_rk.empty else 0
        total_val_cell.fill = orange_fill
        total_val_cell.font = black_font_bold
        total_val_cell.border = thin_border
        total_val_cell.number_format = 'Rp #,##0'
        
        # Add Signatures
        sig_start_row = total_row_idx + 3
        
        ws[f'B{sig_start_row}'] = "Mengetahui"
        ws[f'A{sig_start_row+6}'] = "PT CHIJUN SMART FREIGHT"
        ws[f'A{sig_start_row+6}'].font = Font(bold=True)
        
        ws[f'G{sig_start_row}'] = "Mengetahui"
        ws[f'G{sig_start_row+6}'] = "PT GLOBAL JET EXPRESS"
        ws[f'G{sig_start_row+6}'].font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 18

        
        
    output.seek(0)
    
    # Generate Filename
    # Pattern: "22-31 Desember 2025_BGR_CSF_REPORT W4" -> "INVOICE GLOBAL JET EXPRESS-BGR 22-31 DESEMBER 2025"
    export_name = 'Consolidated_Invoice.xlsx'
    if data:
        first_file = data[0].get('source_file', '')
        # Regex to capture Date_Code_
        # Expecting: [Date Part]_[Code]_[Rest]
        # Example: 22-31 Desember 2025_BGR_...
        match = re.match(r'^(.+?)_([A-Za-z0-9]+)_', first_file)
        if match:
            date_part = match.group(1).strip().upper()
            city_code = match.group(2).strip().upper()
            export_name = f"INVOICE GLOBAL JET EXPRESS-{city_code} {date_part}.xlsx"
        else:
             # Fallback if pattern doesn't match: Try to just use the filename prefix?
             # Or just prepend INVOICE GLOBAL JET EXPRESS
             clean_name = first_file.rsplit('.', 1)[0].upper()
             export_name = f"INVOICE {clean_name}.xlsx"

    return send_file(
        output,
        download_name=export_name,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
