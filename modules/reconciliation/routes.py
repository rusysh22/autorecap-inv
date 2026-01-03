import os
import io
import pandas as pd
from flask import Blueprint, render_template, request, jsonify, send_file
from openpyxl import load_workbook
import datetime

reconciliation_bp = Blueprint('reconciliation', __name__, 
                            template_folder='../../templates/reconciliation', 
                            static_folder='../../static/reconciliation')

# Coordinates configuration
COORD_MAP = {
    "tagihan_kepada": "B7",
    "dikirim_ke": "B12",
    "no_invoice": "J8",
    "invoice_date": "J13",
    "currency": "K13",
    "due_date": "K15"
}

def format_value(val):
    """Utility to format values cleanly."""
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()

def safe_float_convert(val):
    """Convert value to float, handling strings like 'Rp 100.000' or regular numbers."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    
    # Clean string: remove Rp, dots (if thousands), replace comma with dot (if decimal)
    s_val = str(val).lower().replace("rp", "").replace(" ", "")
    
    # Try standard float conversion
    try:
        return float(s_val)
    except ValueError:
        pass
        
    # Handle Indonesia thousands separator (.) and decimal (,)
    if "," in s_val and "." in s_val:
            s_val = s_val.replace(".", "").replace(",", ".")
    elif "." in s_val and s_val.count(".") > 1: # Mulitple dots = thousands
            s_val = s_val.replace(".", "")
    elif "," in s_val: # Check if comma is decimal
            s_val = s_val.replace(",", ".")
            
    try:
        return float(s_val)
    except ValueError:
        return 0.0

def find_value_in_col_k(ws, keyword_list, search_col="J", value_col="K", max_row=50):
    """
    Search for a keyword in 'search_col' (default J) and return the value from 'value_col' (default K).
    Returns raw value or None if not found.
    """
    for row in range(1, max_row + 1):
        cell_val = ws[f"{search_col}{row}"].value
        if cell_val:
            str_val = str(cell_val).lower()
            # Check if any keyword matches
            for key in keyword_list:
                if key in str_val:
                    return ws[f"{value_col}{row}"].value
    return None

def find_value_and_label(ws, keyword_list, search_col="J", value_col="K", max_row=50):
    """
    Search for a keyword in 'search_col' and return (value, label_text).
    """
    for row in range(1, max_row + 1):
        cell_val = ws[f"{search_col}{row}"].value
        if cell_val:
            str_val = str(cell_val).lower()
            for key in keyword_list:
                if key in str_val:
                    return ws[f"{value_col}{row}"].value, str_val
    return None, None

def process_single_file(file_storage):
    # Remove extension from filename for display/export
    filename = os.path.splitext(file_storage.filename)[0]
    error_msg = None
    data = {}

    try:
        # Load workbook with data_only=True to get values, not formulas
        wb = load_workbook(file_storage, data_only=True)
        
        # Select Sheet (Case-insensitive)
        sheet_found = False
        for sheet_name in wb.sheetnames:
            if sheet_name.upper() == "INVOICE":
                ws = wb[sheet_name]
                sheet_found = True
                break
        
        if not sheet_found:
            ws = wb.active

        # Extract fixed coordinates
        extracted = {}
        for key, cell_ref in COORD_MAP.items():
            try:
                val = ws[cell_ref].value
                extracted[key] = format_value(val)
            except Exception as e:
                extracted[key] = ""
        
        # Custom cleanup for Currency field
        # Removes "Currency", newlines, and spaces
        if extracted.get("currency"):
                c_val = extracted["currency"].lower().replace("currency", "").replace("\n", "").strip()
                extracted["currency"] = c_val.upper()
        
        # Dynamic Extraction for DPP, PPN, PPH
        
        raw_dpp = find_value_in_col_k(ws, ["total dasar pengenaan pajak","total dasar pengenaan pajak (asli)", "dpp"], search_col="H")
        raw_diskon = find_value_in_col_k(ws, ["total diskon", "diskon"], search_col="H")
        
        # PPN Logic with 'Dibebaskan' check
        raw_ppn, ppn_label = find_value_and_label(ws, ["total ppn (1.1%)", "total ppn", "ppn"], search_col="H")
        if ppn_label and "dibebaskan" in ppn_label:
            raw_ppn = 0
            
        raw_pph = find_value_in_col_k(ws, ["total pph 23 (2%)", "total pph", "pph 23", "pph"], search_col="H")
        
        extracted['dpp'] = format_value(raw_dpp) if raw_dpp is not None else "0"
        extracted['diskon'] = format_value(raw_diskon) if raw_diskon is not None else "0"
        extracted['ppn'] = format_value(raw_ppn) if raw_ppn is not None else "0"
        extracted['pph'] = format_value(raw_pph) if raw_pph is not None else "0"

        val_dpp = safe_float_convert(raw_dpp)
        val_diskon = safe_float_convert(raw_diskon)
        val_ppn = safe_float_convert(raw_ppn)
        val_pph = safe_float_convert(raw_pph)
        
        # Calculate Total Bayar = DPP - Diskon + PPN - PPH
        # Assuming Diskon is a reduction.
        val_total = val_dpp - val_diskon + val_ppn - val_pph
        
        # Store as float/number for good JSON and Excel export
        extracted['dpp'] = val_dpp
        extracted['diskon'] = val_diskon
        extracted['ppn'] = val_ppn
        extracted['pph'] = val_pph
        extracted['total_bayar'] = val_total
        
        # Validation
        # Rules: No Invoice (J8) must not be empty.
        no_inv = extracted.get("no_invoice")
        
        if not no_inv:
            error_msg = f"Validasi Gagal: 'No. Invoice' (J8) tidak ditemukan."
            status = "failed"
        else:
            status = "success"
            data = extracted
            data['filename'] = filename

    except Exception as e:
        status = "failed"
        error_msg = f"Error processing file: {str(e)}"

    return {
        "status": status,
        "filename": filename,
        "data": data,
        "error": error_msg
    }

@reconciliation_bp.route('/')
def index():
    return render_template('rekon_index.html')

@reconciliation_bp.route('/process', methods=['POST'])
def process_files():
    if 'files' not in request.files:
        return jsonify({"error": "No files uploaded"}), 400
    
    files = request.files.getlist('files')
    results = []
    
    for file in files:
        if not file.filename.lower().endswith('.xlsx'):
            results.append({
                "status": "failed",
                "filename": file.filename,
                "error": "Bukan file Excel (.xlsx)"
            })
            continue
            
        res = process_single_file(file)
        results.append(res)
        
    return jsonify(results)

@reconciliation_bp.route('/export', methods=['POST'])
def export_excel():
    json_data = request.json
    if not json_data:
        return jsonify({"error": "No data to export"}), 400

    # Create DataFrame
    col_mapping = {
        "filename": "Filename",
        "tagihan_kepada": "Tagihan Kepada",
        "dikirim_ke": "Dikirim Ke",
        "no_invoice": "No. Invoice",
        "invoice_date": "Invoice Date",
        "currency": "Currency",
        "due_date": "Due Date",
        "dpp": "DPP",
        "diskon": "Total Diskon",
        "ppn": "Total PPN",
        "pph": "Total PPH",
        "total_bayar": "Total Bayar"
    }
    
    df = pd.DataFrame(json_data)
    
    # Rename columns
    df.rename(columns=col_mapping, inplace=True)
    
    # Ensure column order
    desired_order = [
        "Filename", "Tagihan Kepada", "Dikirim Ke", "No. Invoice", 
        "Invoice Date", "Currency", "Due Date", "DPP", "Total Diskon", "Total PPN", 
        "Total PPH", "Total Bayar"
    ]
    
    # Add missing columns if any
    for col in desired_order:
        if col not in df.columns:
            df[col] = ""
            
    df = df[desired_order]
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Rekap Invoice')
    
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Rekap_Invoice_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
